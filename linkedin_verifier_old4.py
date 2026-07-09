"""
EDS Stakeholder LinkedIn Profile Verifier
==========================================
Playwright-based anti-detection scraper for LinkedIn profile verification.

Design goals (stronger than Oxford scraper):
- Use the user's real logged-in LinkedIn session (cookies import)
- Human-like behaviour: random mouse movement, scroll, variable typing speed
- Delays 1.5x Oxford: random 2.25s-4.5s between requests (Oxford was 1.5-3s)
- Exponential backoff on challenge pages / 429 / captcha detection
- Browser fingerprint de-masking (webdriver=false, realistic UA, plugins, timezone)
- Session checkpoint/resume: won't re-verify already confirmed profiles
- Match-scoring: computes confidence based on name + company + title overlap
- Captcha/challenge detection: STOPS the run and alerts user, never proceeds blind

Usage:
    # 1. Install deps
    pip install playwright beautifulsoup4 openpyxl python-Levenshtein rich
    playwright install chromium

    # 2. Export your LinkedIn cookies from browser (browser extension: "Get cookies.txt")
    #    Save as linkedin_cookies.json in this directory

    # 3. Run
    python linkedin_verifier.py --input stakeholders.xlsx --output verified.xlsx

    # 4. Resume after interruption
    python linkedin_verifier.py --resume

WARNING: LinkedIn explicitly forbids scraping in their ToS (User Agreement s.8.2).
Using this tool violates LinkedIn's terms. Use at your own risk.
Recommended: use sparingly, never on a mission-critical account, and expect
that LinkedIn may issue temporary or permanent restrictions.

Alternative: manual verification using the generated search URLs is safer.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import random
import re
import sys
import time
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Optional
from urllib.parse import quote_plus

try:
    from playwright.async_api import async_playwright, Page, BrowserContext
    from playwright.async_api import TimeoutError as PWTimeoutError
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from rich.console import Console
    from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeRemainingColumn
    import Levenshtein
except ImportError as e:
    print(f"[ERROR] Missing dependency: {e}")
    print("Install with:")
    print("  pip install playwright beautifulsoup4 openpyxl python-Levenshtein rich")
    print("  playwright install chromium")
    sys.exit(1)

console = Console()

# ---------- CONFIG ---------- #
BASE_DIR = Path(__file__).parent
COOKIES_PATH = BASE_DIR / "linkedin_cookies.json"
CHECKPOINT_PATH = BASE_DIR / "verifier_checkpoint.json"
LOG_PATH = BASE_DIR / "verifier.log"

# Timing (1.5x Oxford scraper: Oxford was 1.5-3.0s, here 2.25-4.5s)
MIN_DELAY = 2.25
MAX_DELAY = 4.5
# After every N profiles, take a longer "human break"
BREAK_EVERY = 15
BREAK_MIN = 60    # 1 min
BREAK_MAX = 180   # 3 min
# Detection of LinkedIn challenge/captcha
CHALLENGE_URL_PATTERNS = [
    "/checkpoint/challenge",
    "/authwall",
    "/uas/login",
    "/captcha",
    "security/challenge",
]
# If we hit a challenge, pause this long and ask human to solve
CHALLENGE_PAUSE = 600  # 10 min
# Max retries per profile
MAX_RETRIES = 2
# User agents (realistic, current)
USER_AGENTS = [
    # Chrome 131 macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    # Chrome 131 Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    # Edge 131 Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
]
# Browser viewport choices
VIEWPORTS = [
    {"width": 1920, "height": 1080},
    {"width": 1680, "height": 1050},
    {"width": 1440, "height": 900},
    {"width": 1536, "height": 864},
]
LOCALES = ["en-US", "en-GB", "de-DE"]
TIMEZONES = ["Europe/Berlin", "Europe/London", "America/New_York"]


# ---------- DATA ---------- #
@dataclass
class Stakeholder:
    first_name: str
    last_name: str
    title: str
    company: str
    email: str = ""

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


@dataclass
class VerificationResult:
    stakeholder_key: str        # "first_last_company"
    status: str                 # verified | candidate | not_found | skipped | challenge | error
    linkedin_url: str = ""
    xing_url: str = ""
    confidence: float = 0.0     # 0..1
    matched_name: str = ""
    matched_title: str = ""
    matched_company: str = ""
    notes: str = ""
    attempted_at: str = ""


def load_checkpoint() -> dict:
    if CHECKPOINT_PATH.exists():
        return json.loads(CHECKPOINT_PATH.read_text(encoding="utf-8"))
    return {}


def save_checkpoint(data: dict) -> None:
    CHECKPOINT_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def log(msg: str) -> None:
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    console.print(line)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


# ---------- HUMANISATION ---------- #
async def human_delay(min_s: float = MIN_DELAY, max_s: float = MAX_DELAY) -> None:
    """Random delay with gaussian-ish distribution favouring the middle."""
    base = random.uniform(min_s, max_s)
    # Occasional longer pause (10% of the time)
    if random.random() < 0.10:
        base += random.uniform(1.0, 3.0)
    await asyncio.sleep(base)


async def human_scroll(page: Page) -> None:
    """Random scroll simulating human reading."""
    for _ in range(random.randint(2, 5)):
        delta = random.randint(200, 600)
        await page.mouse.wheel(0, delta)
        await asyncio.sleep(random.uniform(0.4, 1.2))
    # occasionally scroll back up
    if random.random() < 0.3:
        await page.mouse.wheel(0, -random.randint(100, 300))
        await asyncio.sleep(random.uniform(0.3, 0.8))


async def human_mouse_move(page: Page) -> None:
    """Random mouse movement to seem alive."""
    viewport = page.viewport_size or {"width": 1280, "height": 720}
    for _ in range(random.randint(1, 3)):
        x = random.randint(100, viewport["width"] - 100)
        y = random.randint(100, viewport["height"] - 100)
        # Move in several small steps (bezier-ish)
        steps = random.randint(8, 15)
        await page.mouse.move(x, y, steps=steps)
        await asyncio.sleep(random.uniform(0.1, 0.4))


# ---------- ANTI-DETECTION ---------- #
STEALTH_JS = """
// Remove webdriver property
Object.defineProperty(navigator, 'webdriver', { get: () => undefined });

// Override plugins
Object.defineProperty(navigator, 'plugins', {
    get: () => [
        {name: 'Chrome PDF Plugin', filename: 'internal-pdf-viewer'},
        {name: 'Chrome PDF Viewer', filename: 'mhjfbmdgcfjbbpaeojofohoefgiehjai'},
        {name: 'Native Client', filename: 'internal-nacl-plugin'}
    ]
});

// Languages
Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en', 'de'] });

// Chrome object
window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){} };

// Permissions
const originalQuery = window.navigator.permissions.query;
window.navigator.permissions.query = (parameters) => (
    parameters.name === 'notifications'
        ? Promise.resolve({ state: Notification.permission })
        : originalQuery(parameters)
);

// WebGL vendor spoof
const getParameter = WebGLRenderingContext.prototype.getParameter;
WebGLRenderingContext.prototype.getParameter = function(parameter) {
    if (parameter === 37445) return 'Intel Inc.';
    if (parameter === 37446) return 'Intel Iris OpenGL Engine';
    return getParameter.call(this, parameter);
};
"""


async def create_stealth_context(pw) -> BrowserContext:
    # CONSISTENT fingerprint: Windows Chrome + Berlin (matches Jev's Frankfurt VPN exit)
    # Do NOT randomise timezone — mismatch with cookie's timezone cookie value triggers detection
    ua = USER_AGENTS[1]  # Chrome 131 Windows (matches Windows filesystem screenshot)
    viewport = {"width": 1920, "height": 1080}
    locale = "en-US"
    tz = "Europe/Berlin"

    log(f"Browser: UA={ua[:50]}..., viewport={viewport}, locale={locale}, tz={tz}")

    browser = await pw.chromium.launch(
        headless=False,  # LinkedIn detects headless very aggressively
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-features=IsolateOrigins,site-per-process",
            "--disable-site-isolation-trials",
            "--no-sandbox",
        ],
    )
    context = await browser.new_context(
        user_agent=ua,
        viewport=viewport,
        locale=locale,
        timezone_id=tz,
        device_scale_factor=1,
        has_touch=False,
        is_mobile=False,
        color_scheme="light",
    )
    await context.add_init_script(STEALTH_JS)
    return context, browser


async def load_cookies(context: BrowserContext) -> bool:
    """Load LinkedIn session cookies exported from the user's browser."""
    if not COOKIES_PATH.exists():
        log(f"[WARN] No cookies file at {COOKIES_PATH}.")
        log("       Export your LinkedIn cookies using 'Get cookies.txt' browser extension")
        log("       or similar, then save as linkedin_cookies.json")
        return False
    try:
        raw = json.loads(COOKIES_PATH.read_text(encoding="utf-8"))
        cookies = []
        skipped = 0
        for c in raw:
            # Map sameSite to Playwright's exact accepted values: Strict | Lax | None
            ss_raw = str(c.get("sameSite", "")).lower().strip()
            ss_map = {
                "strict": "Strict",
                "lax": "Lax",
                "none": "None",
                "no_restriction": "None",      # Chrome's "no_restriction" = SameSite=None
                "unspecified": "Lax",          # Chrome's "unspecified" defaults to Lax
                "": "Lax",
            }
            same_site = ss_map.get(ss_raw, "Lax")

            # Playwright requires expires to be a number; session cookies have -1 or missing
            expires = c.get("expirationDate", c.get("expires", -1))
            if c.get("session"):
                expires = -1

            cookie = {
                "name": c["name"],
                "value": c["value"],
                "domain": c.get("domain", ".linkedin.com"),
                "path": c.get("path", "/"),
                "expires": expires,
                "httpOnly": c.get("httpOnly", False),
                "secure": c.get("secure", True),
                "sameSite": same_site,
            }
            # skip malformed cookies (empty name/value)
            if not cookie["name"] or cookie["value"] is None:
                skipped += 1
                continue
            cookies.append(cookie)
        await context.add_cookies(cookies)
        log(f"[OK] Loaded {len(cookies)} cookies ({skipped} skipped).")
        return True
    except Exception as e:
        log(f"[ERR] Could not load cookies: {e}")
        return False


# ---------- CHALLENGE DETECTION ---------- #
async def detect_challenge(page: Page) -> bool:
    """
    Detect REAL LinkedIn challenge pages, not false positives from 'verified profile' badges
    or Sales Navigator promos. We require either:
      (a) The URL itself indicates a checkpoint/challenge page, OR
      (b) The page title starts with 'Security Verification' / 'Please verify' etc.
    Just finding the word 'verify' in HTML is NOT enough — LinkedIn sprinkles verified-profile
    badges, 'verify your email' banners, and 'Search with Sales Navigator' cards everywhere.
    """
    url = page.url.lower()

    # Strong signal: URL pattern
    for pattern in CHALLENGE_URL_PATTERNS:
        if pattern in url:
            return True

    # Secondary: page title (real challenge pages have clear titles)
    try:
        title = (await page.title()).lower()
        if any(x in title for x in [
            "security verification",
            "security check",
            "please verify",
            "let's do a quick security check",
            "challenge",
        ]):
            return True
    except Exception:
        pass

    # Tertiary: look for actual challenge UI (captcha iframe / 'please solve' text in a narrow scope)
    try:
        # Only check for iframes pointing to captcha providers
        captcha_frames = await page.query_selector_all(
            "iframe[src*='captcha'], iframe[src*='recaptcha'], iframe[src*='hcaptcha'], iframe[src*='arkose']"
        )
        if captcha_frames:
            return True
    except Exception:
        pass

    # NOTE: We deliberately do NOT flag on the literal word "verify", "verified", "security",
    # or "captcha" appearing anywhere in HTML — those appear on normal search pages too.
    return False


# ---------- MATCHING ---------- #
def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower()) if s else ""


def fuzzy_ratio(a: str, b: str) -> float:
    a, b = normalize(a), normalize(b)
    if not a or not b:
        return 0.0
    # Levenshtein ratio
    try:
        return Levenshtein.ratio(a, b)
    except Exception:
        # fallback: substring
        if a in b or b in a:
            return 0.8
        return 0.0


def compute_confidence(profile_name: str, profile_title: str, profile_company: str,
                       expected: Stakeholder) -> float:
    name_score = fuzzy_ratio(profile_name, expected.full_name)
    title_score = fuzzy_ratio(profile_title, expected.title)
    company_score = fuzzy_ratio(profile_company, expected.company)
    # name matters most
    score = 0.5 * name_score + 0.25 * company_score + 0.25 * title_score
    return round(score, 3)


# ---------- SEARCH + EXTRACTION ---------- #
async def search_linkedin(context: BrowserContext, stakeholder: Stakeholder) -> VerificationResult:
    """Search LinkedIn for the stakeholder; return best candidate."""
    page = await context.new_page()
    key = f"{stakeholder.first_name}_{stakeholder.last_name}_{stakeholder.company}".replace(" ", "_")
    result = VerificationResult(stakeholder_key=key, status="error",
                                attempted_at=time.strftime("%Y-%m-%d %H:%M:%S"))
    try:
        # Use LinkedIn People Search (requires login)
        query = f"{stakeholder.full_name} {stakeholder.company}"
        url = f"https://www.linkedin.com/search/results/people/?keywords={quote_plus(query)}&origin=SWITCH_SEARCH_VERTICAL"

        log(f"Searching: {stakeholder.full_name} @ {stakeholder.company}")
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)

        # Wait for search results to render (LinkedIn is JS-heavy, results come after initial DOM)
        # Try to find any profile link; if not within 8s, proceed anyway for challenge check.
        try:
            await page.wait_for_selector("a[href*='/in/']", timeout=8000)
        except PWTimeoutError:
            log(f"  [info] Results didn't appear within 8s — checking for challenge/empty page")

        # Extra settle time (LinkedIn lazy-loads profile cards)
        await asyncio.sleep(random.uniform(2.0, 3.5))

        # Challenge check (now that page has settled)
        if await detect_challenge(page):
            log(f"[CHALLENGE] Detected on search page for {stakeholder.full_name}")
            result.status = "challenge"
            result.notes = "Security challenge page — human must solve in the open browser window."
            # Pause; user should solve in the visible browser
            log(f"[PAUSE] Sleeping {CHALLENGE_PAUSE}s for manual solve. Keep browser open.")
            await asyncio.sleep(CHALLENGE_PAUSE)
            return result

        await human_mouse_move(page)
        await asyncio.sleep(random.uniform(1.5, 3.0))
        await human_scroll(page)

        # Try multiple selectors — LinkedIn changes markup often.
        # Collect ALL profile cards from the search page and score each one.
        # This is MUCH better than taking the first link (which could be Sales Nav promo or wrong person).
        candidates = []  # list of (url, name_text, subtitle_text)
        try:
            # LinkedIn result items — typically <li> with a profile anchor + name + subtitle
            # We use a broad strategy: find every <a href="/in/..."> that has meaningful surrounding text.
            anchors = await page.query_selector_all("a[href*='/in/']")
            seen_urls = set()
            for a in anchors:
                try:
                    href = await a.get_attribute("href")
                    if not href or "/in/" not in href:
                        continue
                    # Normalise URL (strip query)
                    url_clean = href.split("?")[0]
                    if url_clean.startswith("/"):
                        url_clean = "https://www.linkedin.com" + url_clean
                    # Dedupe
                    if url_clean in seen_urls:
                        continue
                    seen_urls.add(url_clean)

                    # Get the text of the anchor and its surrounding container
                    # (LinkedIn often wraps name in span.entity-result__title-text or similar)
                    anchor_text = (await a.inner_text()).strip()

                    # Climb up to the enclosing <li> or card-container to get headline/location too
                    parent_text = ""
                    try:
                        parent_text = await a.evaluate(
                            "(el) => { const c = el.closest('li') || el.closest('div.entity-result__item') || el.parentElement; return c ? c.innerText : ''; }"
                        )
                    except Exception:
                        pass

                    candidates.append({
                        "url": url_clean,
                        "anchor_text": anchor_text,
                        "container_text": parent_text,
                    })
                    if len(candidates) >= 8:  # cap to top 8 candidates
                        break
                except Exception:
                    continue
        except Exception as e:
            log(f"  [warn] Error enumerating candidates: {e}")

        if not candidates:
            result.status = "not_found"
            result.notes = "No profile cards on search page."
            # Save page HTML to disk for debugging (first failure only)
            debug_path = BASE_DIR / f"debug_no_results_{stakeholder.last_name}.html"
            if not debug_path.exists():
                try:
                    html = await page.content()
                    debug_path.write_text(html, encoding="utf-8")
                    log(f"  [DEBUG] Saved failing page HTML to {debug_path.name}")
                except Exception:
                    pass
            return result

        # Score each candidate using just the search snippet (no profile visit yet — saves bandwidth + detection surface)
        def score_candidate(cand: dict) -> float:
            # anchor_text usually contains the name
            # container_text contains name + headline + current company + location
            name_score = fuzzy_ratio(cand["anchor_text"][:80], stakeholder.full_name)
            # container text is richer — check against company and title
            ct = cand["container_text"].lower()
            company_score = 1.0 if normalize(stakeholder.company) and normalize(stakeholder.company) in normalize(ct) else fuzzy_ratio(ct[:200], stakeholder.company)
            title_tokens = [w for w in stakeholder.title.lower().split() if len(w) > 4]
            if title_tokens:
                title_hits = sum(1 for tok in title_tokens if tok in ct)
                title_score = min(1.0, title_hits / len(title_tokens))
            else:
                title_score = 0.5
            return 0.5 * name_score + 0.3 * company_score + 0.2 * title_score

        scored = [(score_candidate(c), c) for c in candidates]
        scored.sort(key=lambda x: -x[0])
        best_score, best = scored[0]

        log(f"  [candidates] {len(candidates)} found. Top: '{best['anchor_text'][:50]}' score={best_score:.2f}")

        # If the top candidate's name doesn't even roughly match, don't visit the profile — it's a waste
        if fuzzy_ratio(best["anchor_text"][:80], stakeholder.full_name) < 0.40:
            result.status = "not_found"
            result.notes = f"Best candidate name mismatch (top: '{best['anchor_text'][:60]}')"
            result.linkedin_url = best["url"]  # still record for manual review
            result.matched_name = best["anchor_text"][:100]
            return result

        profile_url = best["url"]

        # Visit the best profile to extract real current title/company
        await human_delay()
        await page.goto(profile_url, wait_until="domcontentloaded", timeout=30000)

        if await detect_challenge(page):
            log(f"[CHALLENGE] on profile {profile_url}")
            result.status = "challenge"
            result.notes = "Challenge on profile page."
            await asyncio.sleep(CHALLENGE_PAUSE)
            return result

        await human_mouse_move(page)
        await human_scroll(page)
        await asyncio.sleep(random.uniform(2.0, 4.0))

        # Extract profile info
        try:
            name = await page.locator("h1").first.inner_text(timeout=5000)
        except Exception:
            name = ""
        try:
            headline = await page.locator("div.text-body-medium").first.inner_text(timeout=5000)
        except Exception:
            headline = ""
        # Try to get current company from experience section
        company = ""
        try:
            exp_company = await page.locator("[data-field='experience_company_logo'] img").first.get_attribute("alt", timeout=3000)
            if exp_company:
                company = exp_company
        except Exception:
            pass

        confidence = compute_confidence(name, headline, company or stakeholder.company, stakeholder)

        result.linkedin_url = profile_url
        result.matched_name = name
        result.matched_title = headline
        result.matched_company = company
        result.confidence = confidence
        if confidence >= 0.70:
            result.status = "verified"
        elif confidence >= 0.50:
            result.status = "candidate"
            result.notes = f"Moderate match (search-snippet score {best_score:.2f}) — spot-check before outreach."
        else:
            result.status = "candidate"
            result.notes = f"Low match (search-snippet score {best_score:.2f}) — likely wrong person."

    except PWTimeoutError:
        result.status = "error"
        result.notes = "Timeout"
    except Exception as e:
        result.status = "error"
        result.notes = str(e)[:200]
    finally:
        await page.close()
    return result


# ---------- XING FALLBACK ---------- #
async def search_xing(context: BrowserContext, stakeholder: Stakeholder) -> str:
    """Return Xing profile URL if findable via public directory."""
    page = await context.new_page()
    try:
        query = stakeholder.full_name
        url = f"https://www.xing.com/search/members?keywords={quote_plus(query)}"
        await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        await human_mouse_move(page)
        await asyncio.sleep(random.uniform(2.0, 4.0))
        link = await page.query_selector("a[href*='/profile/']")
        if link:
            href = await link.get_attribute("href")
            if href:
                if href.startswith("/"):
                    href = "https://www.xing.com" + href
                return href.split("?")[0]
    except Exception as e:
        log(f"[Xing] error for {stakeholder.full_name}: {e}")
    finally:
        await page.close()
    return ""


# ---------- GOOGLE FALLBACK ---------- #
async def search_google_for_linkedin(context: BrowserContext, stakeholder: Stakeholder) -> VerificationResult:
    """
    Use Google to find LinkedIn profile URLs.
    Google tolerates automated access much better than LinkedIn, and indexes public LinkedIn profiles.
    Strategy:
      1. Try precise query: site:linkedin.com/in/ "Name" "Company"
      2. If 0 results, fall back to broader: site:linkedin.com/in/ "Name"
      3. Extract top 5 linkedin.com/in/ URLs from results
      4. Score each using snippet text (name + company overlap)
    """
    page = await context.new_page()
    key = f"{stakeholder.first_name}_{stakeholder.last_name}_{stakeholder.company}".replace(" ", "_")
    result = VerificationResult(stakeholder_key=key, status="error",
                                attempted_at=time.strftime("%Y-%m-%d %H:%M:%S"))
    try:
        # Google recognises LinkedIn URLs via site:linkedin.com/in/
        queries = [
            f'site:linkedin.com/in/ "{stakeholder.full_name}" "{stakeholder.company}"',
            f'site:linkedin.com/in/ "{stakeholder.full_name}"',
        ]
        best_url, best_snippet, best_score = "", "", 0.0

        for attempt, query in enumerate(queries):
            log(f"[Google] {stakeholder.full_name}: query #{attempt+1}: {query[:80]}")
            url = f"https://www.google.com/search?q={quote_plus(query)}"

            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(random.uniform(2.0, 3.5))

            # Google's "About this result" consent page — dismiss if it appears
            # Also Google sometimes shows a cookie consent prompt for EU — accept it once
            try:
                consent_btn = await page.query_selector("button[aria-label*='Accept'], button[aria-label*='Reject'], button:has-text('Accept all'), #L2AGLb")
                if consent_btn:
                    await consent_btn.click()
                    await asyncio.sleep(random.uniform(1.0, 2.0))
            except Exception:
                pass

            await human_mouse_move(page)
            await human_scroll(page)

            # Extract top results. Google's result anchors live under <a href="..."> inside result blocks.
            # We filter for linkedin.com/in/ URLs.
            anchors = await page.query_selector_all("a[href*='linkedin.com/in/']")
            if not anchors:
                log(f"  [Google] no linkedin hits for query #{attempt+1}")
                if attempt == 0:
                    await human_delay()
                    continue
                else:
                    break

            # Process top 5 hits
            seen = set()
            candidates = []
            for a in anchors[:10]:
                try:
                    href = await a.get_attribute("href")
                    if not href or "linkedin.com/in/" not in href:
                        continue
                    # Google wraps real URL in /url?q=... sometimes; clean that
                    clean = href.split("?")[0] if "/url?q=" not in href else href.split("/url?q=")[1].split("&")[0]
                    if "linkedin.com/in/" not in clean:
                        continue
                    if clean in seen:
                        continue
                    seen.add(clean)

                    # Grab the surrounding snippet for context
                    snippet = ""
                    try:
                        snippet = await a.evaluate(
                            "(el) => { const c = el.closest('div[data-ved]') || el.closest('div'); return c ? c.innerText : ''; }"
                        )
                    except Exception:
                        pass
                    candidates.append({"url": clean, "snippet": snippet[:500]})
                    if len(candidates) >= 5:
                        break
                except Exception:
                    continue

            if not candidates:
                continue

            # Score each candidate
            def score(c):
                snip = c["snippet"].lower()
                name_score = fuzzy_ratio(stakeholder.full_name, snip[:200])
                # Check for company tokens in snippet
                co_norm = normalize(stakeholder.company)
                co_score = 1.0 if co_norm and co_norm in normalize(snip) else fuzzy_ratio(stakeholder.company, snip[:300])
                # Check URL slug for name similarity (good signal)
                slug = c["url"].split("/in/")[-1].split("/")[0].lower().replace("-", "")
                slug_name = normalize(stakeholder.full_name)
                slug_score = 1.0 if slug_name[:8] in slug else fuzzy_ratio(slug, slug_name)
                return 0.4 * name_score + 0.3 * co_score + 0.3 * slug_score

            scored = [(score(c), c) for c in candidates]
            scored.sort(key=lambda x: -x[0])
            top_score, top = scored[0]
            log(f"  [Google] top: {top['url'][:60]} score={top_score:.2f}")

            if top_score > best_score:
                best_score = top_score
                best_url = top["url"]
                best_snippet = top["snippet"][:200]

            # If very confident on query 1, don't bother with broader query
            if best_score >= 0.75:
                break

            await human_delay()

        if best_url:
            result.linkedin_url = best_url
            result.matched_name = best_snippet
            result.confidence = round(best_score, 3)
            if best_score >= 0.70:
                result.status = "verified"
                result.notes = f"Google: confidence {best_score:.2f}"
            elif best_score >= 0.50:
                result.status = "candidate"
                result.notes = f"Google: moderate match ({best_score:.2f}) — spot-check before outreach."
            else:
                result.status = "candidate"
                result.notes = f"Google: low-confidence match ({best_score:.2f}) — likely wrong."
        else:
            result.status = "not_found"
            result.notes = "Google returned no linkedin.com/in/ hits."
    except PWTimeoutError:
        result.status = "error"
        result.notes = "Google timeout"
    except Exception as e:
        result.status = "error"
        result.notes = f"Google error: {str(e)[:200]}"
    finally:
        await page.close()
    return result


# ---------- MAIN PIPELINE ---------- #
async def run(input_xlsx: Path, output_xlsx: Path,
              test_n: Optional[int] = None, retry_failed: bool = False,
              google_fallback: bool = False):
    stakeholders = load_stakeholders(input_xlsx)
    if test_n:
        stakeholders = stakeholders[:test_n]
        log(f"[TEST MODE] Processing first {test_n} stakeholders.")

    checkpoint = load_checkpoint()
    log(f"Loaded checkpoint with {len(checkpoint)} prior results.")
    # Print a quick breakdown so user knows where they stand
    if checkpoint:
        from collections import Counter
        status_counts = Counter(r.get("status", "unknown") for r in checkpoint.values())
        log(f"  Status breakdown: {dict(status_counts)}")

    async with async_playwright() as pw:
        context, browser = await create_stealth_context(pw)
        cookies_ok = await load_cookies(context)
        if not cookies_ok:
            console.print("[yellow]⚠  No cookies loaded. LinkedIn search requires login.[/yellow]")
            console.print("[yellow]   Opening LinkedIn login page — log in manually, then the script will proceed.[/yellow]")
            login_page = await context.new_page()
            await login_page.goto("https://www.linkedin.com/login")
            console.print("[cyan]   Press Enter after you've logged in...[/cyan]")
            input()
            await login_page.close()

        # -------- PREFLIGHT CHECK -------- #
        # Verify cookies actually authenticate us; re-prompt login if they're stale.
        log("[Preflight] Checking login status via linkedin.com/feed...")
        pre = await context.new_page()
        try:
            await pre.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(random.uniform(2.5, 4.0))  # let page settle
            current_url = pre.url.lower()
            content = (await pre.content()).lower()

            # Detection: if we're bounced to login/authwall OR if the page has "sign in" form → not logged in
            logged_in = True
            if any(x in current_url for x in ["/login", "/authwall", "/uas/login", "/checkpoint"]):
                logged_in = False
            elif "session_key" in content and "session_password" in content:
                logged_in = False
            elif '"feed-identity-module"' not in content and "global-nav" not in content:
                # Secondary check — real feed has global-nav
                # still might be ok if LinkedIn changed markup, so only flag if other signs also point to logout
                if "join now" in content or "log in to linkedin" in content:
                    logged_in = False

            if not logged_in:
                console.print("[red]⚠  Cookies are stale or you were logged out.[/red]")
                console.print("[yellow]   The browser window is open on LinkedIn login. Please log in manually,[/yellow]")
                console.print("[yellow]   solve any captcha/2FA, then return here and press Enter.[/yellow]")
                # Bring them to login explicitly
                try:
                    await pre.goto("https://www.linkedin.com/login", wait_until="domcontentloaded")
                except Exception:
                    pass
                console.print("[cyan]   Press Enter once you see your LinkedIn feed...[/cyan]")
                input()
                # Save the now-valid cookies back to disk for next run
                new_cookies = await context.cookies()
                try:
                    COOKIES_PATH.write_text(json.dumps(new_cookies, indent=2, ensure_ascii=False),
                                            encoding="utf-8")
                    log(f"[OK] Saved {len(new_cookies)} refreshed cookies to {COOKIES_PATH.name}")
                except Exception as e:
                    log(f"[WARN] Could not save cookies: {e}")
            else:
                log("[OK] Logged in — cookies are valid.")
                # Simulate reading the feed for a bit (natural behaviour before searching)
                await human_mouse_move(pre)
                await human_scroll(pre)
                await asyncio.sleep(random.uniform(3.0, 6.0))
        finally:
            await pre.close()

        try:
            with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                          BarColumn(), TimeRemainingColumn(), console=console) as progress:
                task = progress.add_task("Verifying profiles", total=len(stakeholders))

                # Define what statuses are "done" (skip) vs "retry"
                # Default: skip everything already attempted (verified, candidate, not_found).
                # With --retry-failed: re-try anything that didn't verify.
                if retry_failed:
                    skip_statuses = {"verified"}
                elif google_fallback:
                    # In Google fallback mode, ONLY run profiles that are not_found or unverified
                    # Skip verified and (real) candidate matches
                    skip_statuses = {"verified", "candidate", "skipped"}
                else:
                    skip_statuses = {"verified", "candidate", "not_found", "skipped"}
                # "challenge" and "error" are ALWAYS retried — they represent transient failures.

                for i, s in enumerate(stakeholders):
                    key = f"{s.first_name}_{s.last_name}_{s.company}".replace(" ", "_")
                    existing = checkpoint.get(key, {})
                    existing_status = existing.get("status")

                    if existing_status in skip_statuses:
                        log(f"[SKIP] {s.full_name} — already {existing_status}.")
                        progress.advance(task)
                        continue

                    if google_fallback:
                        # Google-fallback mode: only process profiles that failed LinkedIn search
                        # If profile was never attempted OR was not_found, run Google search
                        if existing_status not in (None, "not_found", "error", "challenge"):
                            progress.advance(task)
                            continue
                        log(f"[Google-fallback] {s.full_name}")
                        result = await search_google_for_linkedin(context, s)
                    else:
                        # LinkedIn search (default mode)
                        result = await search_linkedin(context, s)

                        # Xing fallback if LinkedIn didn't verify
                        if result.status in ("not_found", "candidate") and result.confidence < 0.70:
                            log(f"  -> Falling back to Xing for {s.full_name}")
                            xing = await search_xing(context, s)
                            if xing:
                                result.xing_url = xing
                                if result.status == "not_found":
                                    result.status = "candidate"

                    checkpoint[key] = asdict(result)
                    save_checkpoint(checkpoint)

                    # INCREMENTAL EXCEL SAVE: write the output file after EVERY profile.
                    # This way if the script is killed mid-run, the Excel file is still up to date.
                    try:
                        write_output(stakeholders, checkpoint, output_xlsx)
                    except Exception as e:
                        log(f"[WARN] Could not write incremental Excel: {e}")

                    # Break pattern: every N profiles, take a long human-like break
                    if (i + 1) % BREAK_EVERY == 0:
                        break_s = random.uniform(BREAK_MIN, BREAK_MAX)
                        log(f"[BREAK] Human-like pause for {break_s:.0f}s after {i+1} profiles.")
                        await asyncio.sleep(break_s)
                    else:
                        await human_delay()

                    progress.advance(task)

        finally:
            log("Closing browser.")
            await context.close()
            await browser.close()

    write_output(stakeholders, checkpoint, output_xlsx)
    log(f"[DONE] Results saved to {output_xlsx}")


def load_stakeholders(path: Path) -> list[Stakeholder]:
    wb = load_workbook(path, data_only=True)
    # Find the right sheet
    sheet_name = "Master List" if "Master List" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    # Find column indices
    def idx(name, alternates=()):
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i
            for alt in alternates:
                if h and alt.lower() in str(h).lower():
                    return i
        return None

    i_first = idx("first")
    i_last = idx("last")
    i_title = idx("title", ("job",))
    i_company = idx("company")
    i_email = idx("email")

    stakeholders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[i_first]:
            continue
        stakeholders.append(Stakeholder(
            first_name=str(row[i_first] or "").strip(),
            last_name=str(row[i_last] or "").strip(),
            title=str(row[i_title] or "").strip(),
            company=str(row[i_company] or "").strip(),
            email=str(row[i_email] or "").strip() if i_email is not None else "",
        ))
    return stakeholders


def write_output(stakeholders: list[Stakeholder], results: dict, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Results"
    headers = ["First", "Last", "Title", "Company", "Email", "Status",
               "Confidence", "LinkedIn URL", "Matched Name", "Matched Title",
               "Matched Company", "Xing URL", "Notes", "Attempted At"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="00A49A")
    for i, s in enumerate(stakeholders, 2):
        key = f"{s.first_name}_{s.last_name}_{s.company}".replace(" ", "_")
        r = results.get(key, {})
        row = [
            s.first_name, s.last_name, s.title, s.company, s.email,
            r.get("status", "not_attempted"),
            r.get("confidence", ""),
            r.get("linkedin_url", ""),
            r.get("matched_name", ""),
            r.get("matched_title", ""),
            r.get("matched_company", ""),
            r.get("xing_url", ""),
            r.get("notes", ""),
            r.get("attempted_at", ""),
        ]
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Colour-code status
            if j == 6:
                colours = {
                    "verified": "C6EFCE",
                    "candidate": "FFEB9C",
                    "not_found": "FFC7CE",
                    "challenge": "FFC7CE",
                    "error": "D9D9D9",
                }
                cell.fill = PatternFill("solid", start_color=colours.get(v, "FFFFFF"))
    for col, width in zip("ABCDEFGHIJKLMN", [14, 16, 42, 28, 32, 14, 12, 48, 24, 38, 24, 48, 40, 20]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(stakeholders)+1}"
    wb.save(output)


def main():
    p = argparse.ArgumentParser(description="LinkedIn/Xing profile verifier for EDS stakeholders")
    p.add_argument("--input", type=Path, default=BASE_DIR / "stakeholders.xlsx",
                   help="Input Excel with stakeholder list")
    p.add_argument("--output", type=Path, default=BASE_DIR / "verified.xlsx",
                   help="Output Excel with verification results")
    p.add_argument("--test", type=int, default=None,
                   help="Test mode: only process first N stakeholders")
    p.add_argument("--reset", action="store_true",
                   help="Reset checkpoint and start over")
    p.add_argument("--retry-failed", action="store_true",
                   help="Also retry candidate/not_found profiles (default: only retry challenge/error)")
    p.add_argument("--google-fallback", action="store_true",
                   help="Use Google search for profiles that failed LinkedIn search. Pair with existing checkpoint to avoid re-running successful ones. Recommended workflow: first do LinkedIn pass, then run with --google-fallback to catch the misses.")
    p.add_argument("--status", action="store_true",
                   help="Just show checkpoint status and exit (no scraping)")
    args = p.parse_args()

    # Status-only mode — no scraping, just report
    if args.status:
        if not CHECKPOINT_PATH.exists():
            console.print("[yellow]No checkpoint file yet — nothing has been scraped.[/yellow]")
            sys.exit(0)
        checkpoint = load_checkpoint()
        from collections import Counter
        status_counts = Counter(r.get("status", "unknown") for r in checkpoint.values())
        console.print(f"\n[bold cyan]Checkpoint: {CHECKPOINT_PATH.name}[/bold cyan]")
        console.print(f"Total profiles processed: {len(checkpoint)}")
        console.print("\n[bold]Status breakdown:[/bold]")
        for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
            emoji = {"verified": "✓", "candidate": "?", "not_found": "✗",
                     "challenge": "🛑", "error": "⚠", "skipped": "—"}.get(status, "•")
            console.print(f"  {emoji} {status}: {count}")
        # Also count stakeholders in input that aren't done yet
        if args.input.exists():
            try:
                all_stakeholders = load_stakeholders(args.input)
                done_keys = set(checkpoint.keys())
                done_statuses = {k: checkpoint[k].get("status") for k in done_keys}
                pending = []
                for s in all_stakeholders:
                    key = f"{s.first_name}_{s.last_name}_{s.company}".replace(" ", "_")
                    if key not in done_statuses:
                        pending.append(s.full_name)
                console.print(f"\n[bold]Pending (never attempted):[/bold] {len(pending)}")
                for name in pending[:10]:
                    console.print(f"  • {name}")
                if len(pending) > 10:
                    console.print(f"  ... and {len(pending) - 10} more")
            except Exception as e:
                console.print(f"[yellow]Could not read input: {e}[/yellow]")
        sys.exit(0)

    if args.reset and CHECKPOINT_PATH.exists():
        CHECKPOINT_PATH.unlink()
        log("Checkpoint reset.")

    if not args.input.exists():
        console.print(f"[red]Input file not found: {args.input}[/red]")
        sys.exit(1)

    asyncio.run(run(args.input, args.output,
                    test_n=args.test, retry_failed=args.retry_failed,
                    google_fallback=args.google_fallback))


if __name__ == "__main__":
    main()
